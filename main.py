import requests
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import time
from multiprocessing.pool import ThreadPool as Pool

def get_all_universities():
    URL = "http://admlist.ru"
    page = requests.get(URL)
    page.encoding = "utf-8"
    page = page.text

    bs = BS(page, "lxml")
    universitys_table_body = bs.find_all("table")[-1].tbody
    anchors = universitys_table_body.find_all("a", recursive=True)
    for a in anchors:
        yield a.text.strip(), URL + "/" + a.get("href")


def get_all_programs(university_link):
    page = requests.get(university_link)
    page.encoding = "utf-8"
    page = page.text

    bs = BS(page, "lxml")
    programs_table_body = bs.find_all("table")[-1].tbody
    anchors = programs_table_body.find_all("a", recursive=True)
    for a in anchors:
        full_program, *direction = a.text.split(",")
        *program_name, slug = full_program.split(" ")
        program_name = " ".join(program_name)
        slug = slug[1:-1]
        short_link = a.get("href")
        *university_app_link, _ = university_link.split("/")
        full_link = "/".join(university_app_link) + "/" + short_link
        yield program_name.strip(), slug.strip(), direction[0].strip() if direction else None, full_link


def parse_program(program_link):
    page = requests.get(program_link)
    page.encoding = "utf-8"
    page = page.text

    bs = BS(page, "lxml")
    table = bs.find_all("table")[-1]
    headers = table.thead.tr.find_all("th")[3:-2]
    yield [th.text.strip() for th in headers]
    for row in table.tbody.find_all("tr"):
        tds = row.find_all("td")
        yield list(map(lambda x: x.text.strip(), tds[3:-2]))


pool = Pool(5)
start = time.time()

def export(uname, pname, slug, direction, plink):
    print(f"Parsing {uname}.{pname}({slug}).{direction}", end=": ")
    try:
        direction = direction if direction is not None else pname
        parsed = parse_program(plink)
        headers = next(parsed)
        wb = Workbook()
        ws = wb.active
        for idx, header in enumerate(headers):
            ws.cell(row=1, column=idx + 1, value=header)
        for row, (snils, att, app_type, *exams) in enumerate(parsed):
            ws.cell(row=row + 2, column=1, value=snils)
            ws.cell(row=row + 2, column=2, value=att)
            ws.cell(row=row + 2, column=3, value=app_type)
            for col, result in enumerate(exams):
                ws.cell(row + 2, column=col + 4, value=int(result))
        wb.save(f"{uname}.{slug}.{direction}.xlsx")
        print("parsed.")
    except:
        print("error")

for uname, ulink in get_all_universities():
    for pname, slug, direction, plink in get_all_programs(ulink):
        pool.apply_async(export, (uname, pname, slug, direction, plink))
            
pool.close()
pool.join()
            
